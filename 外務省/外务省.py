from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
import re
from typing import List, Dict
from openpyxl import Workbook, load_workbook
import os
# 日期转换函数
def convert_japanese_date(japanese_date_str):
    reiwa_match = re.search(r"令和(\d+)年(\d+)月(\d+)日", japanese_date_str)
    if reiwa_match:
        year = int(reiwa_match.group(1)) + 2018  # 令和元年是2019年
        month = reiwa_match.group(2)
        day = reiwa_match.group(3)
        return f"{year}/{month}/{day}"
    else:
        return japanese_date_str


# 提取标题和日期的函数
def extract_title_and_date(driver):
    try:
        title_element = driver.find_element(By.CSS_SELECTOR, "h2.title2")
        date_element = None
        if title_element:
            title_element = driver.find_element(By.CSS_SELECTOR, "main > article > div > h2 > span")
            date_element = driver.find_element(By.CSS_SELECTOR, "main > article > div > h3 > span")
        else:
            title_element = driver.find_element(By.CSS_SELECTOR,
                                                "div:nth-of-type(3) > div:nth-of-type(4) > div:nth-of-type(1) > h2#title2")
            date_element = driver.find_element(By.CSS_SELECTOR, "div:nth-of-type(3) > div:nth-of-type(4) > h3#subtitle")

        if title_element and date_element:
            title = title_element.text.strip()
            date_text = date_element.text.strip()
            combined_text = f"{title} {date_text}"
            print(f"Combined Title and Date: {combined_text}")

            # 提取并转换日期
            converted_date = convert_japanese_date(date_text)
            print(f"Converted Date: {converted_date}")
            return combined_text, converted_date
        else:
            print("未找到标题或日期元素。")
    except Exception as e:
        print("提取标题和日期时出错:", e)
#提取媒体
def extract_media(speaker_info: str) -> str:
    """Extract media name from the speaker section if present."""
    media_match = re.match(r"(.*?)　.*?記者", speaker_info)#r"(.*?)大臣|"|r"(.*?)報道官|"r"(.*?)副大臣|"
    return media_match.group(1).strip() if media_match else ""

#对话处理
def parse_conversation_details(conversation_segments: List[str]) -> List[Dict[str, str]]:
    parsed_conversations = []
    last_topic = ""  # 用于保存上一个对话的主题

    for segment in conversation_segments:
        segment = segment.replace('\n', '')  # 将换行符替换为空格

        # Identify topic and reporter section
        topic_match = re.match(r'(.+?)【([^】]*記者[^】]*)】', segment)
        if topic_match:
            topic, reporter_info = topic_match.groups()
            media_name = extract_media(reporter_info)
            last_topic = topic.strip()  # 更新上一个主题

            # Find question after the reporter section to the next 【 or end of text
            question_start_idx = topic_match.end()
            question_match = re.search(r'([\s\S]*?)(?=【|\Z)', segment[question_start_idx:], flags=re.DOTALL)
            if question_match:
                question_text = question_match.group(1).strip()
                question_end_idx = question_start_idx + question_match.end()

                # Find speaker and response after the question section
                speaker_match = re.search(r'【([^】]*[^記者])】', segment[question_end_idx:])
                if speaker_match:
                    speaker_name = speaker_match.group(1).strip()
                    response_start_idx = question_end_idx + speaker_match.end()
                    response_text = segment[response_start_idx:].strip()  # Remaining text as response

                    # Compile conversation details
                    conversation = {
                        "主题": last_topic,
                        "媒体": media_name.strip(),
                        "提问": question_text.strip(),
                        "发言人": speaker_name.strip(),
                        "回答": response_text.strip(),
                    }
                    parsed_conversations.append(conversation)

        else:
            # 如果没有找到新的主题，继承上一个主题
            reporter_match = re.match(r'【([^】]*記者[^】]*)】', segment)
            if reporter_match:
                reporter_info = reporter_match.group(1)
                media_name = extract_media(reporter_info)

                # Find question after the reporter section to the next 【 or end of text
                question_start_idx = reporter_match.end()
                question_match = re.search(r'([\s\S]*?)(?=【|\Z)', segment[question_start_idx:], flags=re.DOTALL)
                if question_match:
                    question_text = question_match.group(1).strip()
                    question_end_idx = question_start_idx + question_match.end()

                    # Find speaker and response after the question section
                    speaker_match = re.search(r'【([^】]*[^記者])】', segment[question_end_idx:])
                    if speaker_match:
                        speaker_name = speaker_match.group(1).strip()
                        response_start_idx = question_end_idx + speaker_match.end()
                        response_text = segment[response_start_idx:].strip()  # Remaining text as response

                        # Compile conversation details
                        conversation = {
                            "主题": last_topic,
                            "媒体": media_name.strip(),
                            "提问": question_text.strip(),
                            "发言人": speaker_name.strip(),
                            "回答": response_text.strip(),
                        }
                        parsed_conversations.append(conversation)

    return parsed_conversations
#文本分割
def split_conversation_segments(full_text: str) -> List[str]:
    conversation_segments = []
    segment_start_idx = 0  # Start index for each conversation segment

    # Locate all reporter segments as markers for conversation beginnings
    reporter_matches = list(re.finditer(r'【[^】]*記者[^】]*】', full_text))

    for reporter_match in reporter_matches:
        # Get the start index of the current reporter match
        reporter_start_idx = reporter_match.start()

        # Find the preceding period to mark the end of the previous segment
        preceding_period_match = re.search(r'。', full_text[:reporter_start_idx][::-1])

        if preceding_period_match:
            # Calculate the true end index in the text
            segment_end_idx = reporter_start_idx - preceding_period_match.start()
            conversation_segments.append(full_text[segment_start_idx:segment_end_idx].strip())
            # Update start index for the next segment
            segment_start_idx = segment_end_idx

    # Add the final segment from the last start index to the end of the text
    if segment_start_idx < len(full_text):
        conversation_segments.append(full_text[segment_start_idx:].strip())

    return conversation_segments

#提取maincontents内容，并做初步处理
def extract_main_contents(cleaned_text):
    q_and_a_text = ""
    extracted_themes = []  # 保存冒頭発言的主题和内容

    main_contents_text = cleaned_text
    # 提取冒頭发言的内容
    q_and_a_text = main_contents_text
    extracted_text = ""  # 初始化 extracted_text 以防止未定义的引用
    if main_contents_text.startswith("冒頭"):
        print("\n=== Extracting 冒頭発言 ===")
        last_content_end = 0
        if "（1）" in main_contents_text:
            # 处理多主题的情况，将冒頭发言的内容认定为最后一个主题的发言内容的结束
            themes = re.split(r'（\d+）', main_contents_text)
            for idx, theme_content in enumerate(themes):
                if idx == 0 and "冒頭発言" in theme_content:
                    # 提取冒頭发言的主题和内容
                    theme_start = theme_content.find("冒頭発言") + len("冒頭発言")
                    theme_end = theme_content.find("【", theme_start)
                    theme = theme_content[theme_start:theme_end].strip()
                    print(f"冒頭主题: {theme}")
                elif idx > 0:
                    # 多主题处理
                    sub_theme_end = theme_content.find("【")
                    sub_theme = theme_content[:sub_theme_end].strip()
                    if sub_theme:
                        print(f"主题 {idx}: {sub_theme}")

                    # 提取发言内容
                    content_start = theme_content.find("】") + 1
                    next_bracket = theme_content.find("【", content_start)
                    content_end = theme_content.rfind("。", content_start, next_bracket) + 1 if next_bracket != -1 else len(theme_content)
                    content = theme_content[content_start:content_end].strip()
                    print(f"发言内容: {content}")
                    extracted_text += theme_content[:content_end]
                    last_content_end = main_contents_text.find(content) + len(content)

                    # 提取发言人
                    speaker_pattern = re.compile(r"【(.*?)】")
                    speaker_match = speaker_pattern.search(theme_content)
                    speaker = speaker_match.group(1) if speaker_match else "未知"
                    print(f"发言人: {speaker}")
                    print("\n")

                    # 保存冒頭发言的内容
                    extracted_themes.append({
                            "主题": sub_theme,
                            "发言人": speaker,
                            "回答": content
                        })
        else:
            # 单主题处理
            theme_start = main_contents_text.find("冒頭発言") + len("冒頭発言") if "冒頭発言" in main_contents_text else main_contents_text.find("冒頭") + len("冒頭")
            theme_end = main_contents_text.find("【", theme_start)
            theme = main_contents_text[theme_start:theme_end].strip()
            print(f"冒頭主题: {theme}")

            # 提取发言人和发言内容
            speaker_pattern = re.compile(r"【(.*?)】")
            speakers = speaker_pattern.findall(main_contents_text)

            content_start = main_contents_text.find("】", theme_end) + 1
            content_end = content_start
            while True:
                next_bracket = main_contents_text.find("【", content_end)
                if next_bracket == -1:
                    content_end = main_contents_text.rfind("。", content_start) + 1  # 找到最后一个句号
                    break
                next_speaker = speaker_pattern.search(main_contents_text, next_bracket)
                if next_speaker and "記者" in next_speaker.group(1):
                    content_end = main_contents_text.rfind("。", content_start, next_bracket) + 1
                    break
                content_end = next_bracket + 1
            content = main_contents_text[content_start:content_end].strip()
            print(f"发言内容: {content}")
            extracted_text = main_contents_text[:content_end]
            last_content_end = content_end
            speaker = speakers[0] if speakers else "未知"
            print(f"发言人: {speaker}")
            print("\n")

            # 保存冒頭发言的内容
            extracted_themes.append({
                    "主题": theme,
                    "发言人": speaker,
                    "回答": content
                })

        # 去掉冒頭发言的内容，输出 Q&A
        if last_content_end:
            q_and_a_text = main_contents_text[last_content_end:].strip()
    else:
        # 当第一个【】中不包含“記者”时，也将其视为冒頭发言
        first_bracket = main_contents_text.find("【")
        if first_bracket != -1:
            first_bracket_content = re.search(r"【(.*?)】", main_contents_text[first_bracket:])
            if first_bracket_content and "記者" not in first_bracket_content.group(1):
                print("\n=== Extracting 冒頭発言 ===")
                theme_end = first_bracket
                theme = main_contents_text[:theme_end].strip()
                print(f"冒頭主题: {theme}")

                # 提取发言内容
                content_start = main_contents_text.find("】", theme_end) + 1
                content_end = main_contents_text.find("【", content_start)
                if content_end == -1:
                    content_end = len(main_contents_text)
                content = main_contents_text[content_start:content_end].strip()
                print(f"发言内容: {content}")
                extracted_text = main_contents_text[:content_end]
                last_content_end = content_end
                speaker = first_bracket_content.group(1) if first_bracket_content else "未知"
                print(f"发言人: {speaker}")
                print("\n")
                q_and_a_text = main_contents_text[last_content_end:].strip()

                # 保存冒頭发言的内容
                extracted_themes.append({
                        "主题": theme,
                        "发言人": speaker,
                        "回答": content
                    })

    # 进一步处理 Q&A 部分
    if q_and_a_text:
        conversation_segments = split_conversation_segments(q_and_a_text)
        parsed_conversations = parse_conversation_details(conversation_segments)
        for idx, conversation in enumerate(parsed_conversations, 1):
            print(f"Conversation {idx}:")
            for key, value in conversation.items():
                print(f"{key}: {value}")
            print()

    return q_and_a_text, extracted_themes



# 配置driver
driver = webdriver.Chrome()

#输出操作
# 初始化或加载 Excel 文件
def init_excel(file_name):
    if not os.path.exists(file_name):
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        # 写入表头
        headers = ["序号","链接", "标题", "日期", "媒体", "大主题", "小主题", "提问", "发言人", "回答"]
        ws.append(headers)
        wb.save(file_name)
        print(f"Excel 文件已初始化: {file_name}")
    else:
        print(f"加载已存在的 Excel 文件: {file_name}")

# 写入一行数据到 Excel 文件
def append_to_excel(file_name,number, link, title, date, media, main_topic, sub_topic, question, speaker, answer):
    wb = load_workbook(file_name)
    ws = wb.active
    # 添加一行数据
    ws.append([number,link, title, date, media, main_topic, sub_topic, question, speaker, answer])
    wb.save(file_name)
    print(f"已写入数据到 Excel 文件: {file_name}")

#正式操作
file_name = "output.xlsx"

# 初始化 Excel 文件
init_excel(file_name)
try:
    # 打开目标网站
    url = "https://www.mofa.go.jp/mofaj/press/kaiken/gaisho/index.html"
    driver.get(url)

    for idx in []:#遍历年份
        max_i =  if idx ==  else #遍历月份，可根据年份选择月份
        for i in range(1, max_i + 1):  # a[1] 到 a[max]
            try:
                # 动态生成 XPath
                xpath = f"/html/body/div/main/article/div[2]/div/div/div[2]/div[3]/div/table/tbody/tr[{idx}]/td[2]/a[{i}]"
                # 等待并定位到链接元素
                link_element = WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.XPATH, xpath))
                )

                # 获取链接 URL
                link_url = link_element.get_attribute("href")

                # 点击进入该链接
                link_element.click()

                # 使用 XPath 定位到 <div id="pressconf"> 下的所有 dl 元素
                dl_elements = driver.find_elements(By.XPATH, "//div[@id='pressconf']//dl")

                # 过滤出可见的有效 dl 元素
                valid_dl_elements = [dl for dl in dl_elements if dl.is_displayed()]

                # 获取有效的 dl 元素数量
                total_valid_dl = len(valid_dl_elements)

                # 遍历所有的有效 dl 元素
                for dl_index in range(1, total_valid_dl + 1):
                    try:
                        # 动态定位每个 dl 元素的 dt/a 链接
                        target_xpath = f"//dl[{dl_index}]/dt/a"

                        # 等待并定位到目标链接元素
                        target_element = WebDriverWait(driver, 10).until(
                            EC.presence_of_element_located((By.XPATH, target_xpath))
                        )
                        link_url = target_element.get_attribute("href")  #
                        # 点击目标链接进入目标页面
                        target_element.click()

                        # 等待目标页面加载
                        time.sleep(3)  # 根据需要调整时间

                        WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.TAG_NAME, "h2")))

                        # 调用函数提取标题和日期
                        Combined_Title_And_Date, Converted_Date = extract_title_and_date(driver)

                        # 获取目标页面中的 maincontents 部分
                        maincontents_element = WebDriverWait(driver, 10).until(
                            EC.presence_of_element_located((By.ID, "maincontents"))
                        )

                        # 使用 JavaScript 删除不需要的部分
                        driver.execute_script("""
                            var elements = document.querySelectorAll('.rightalign.other-language, .social-btn-top');
                            elements.forEach(function(element) {
                                element.remove();
                            });
                        """)

                        # 获取清理后的 maincontents 下的文本
                        cleaned_text = maincontents_element.text.strip()

                        # 输出清理后的文本内容
                        q_and_a_text, extracted_themes = extract_main_contents(cleaned_text)
                        parsed_conversations = parse_conversation_details(split_conversation_segments(q_and_a_text))

                        # 保存冒頭发言到 Excel
                        for theme in extracted_themes:
                            append_to_excel(
                                file_name,
                                number="",
                                link=link_url,
                                title=Combined_Title_And_Date,
                                date=Converted_Date,
                                media="",
                                main_topic="",
                                sub_topic=theme["主题"],
                                question="",
                                speaker=theme["发言人"],
                                answer=theme["回答"]
                            )

                        # 分割并解析 Q&A 内容
                        conversation_segments = split_conversation_segments(q_and_a_text)
                        parsed_conversations = parse_conversation_details(conversation_segments)

                        # 保存 Q&A 对话到 Excel
                        for conversation in parsed_conversations:
                            append_to_excel(
                                file_name,
                                number="",
                                link=link_url,
                                title=Combined_Title_And_Date,
                                date=Converted_Date,
                                media=conversation["媒体"],
                                main_topic="",
                                sub_topic=conversation["主题"],
                                question=conversation["提问"],
                                speaker=conversation["发言人"],
                                answer=conversation["回答"]
                            )
                        # 返回到主页面，准备爬取下一个链接
                        driver.back()

                        # 等待主页面加载
                        time.sleep(3)

                    except Exception as e:
                        print(f"Error processing dl[{dl_index}]: {e}")
                        break  # 如果某个链接无法访问，则跳出循环
                driver.get(url)



            except Exception as e:
                print(f"Error processing link {i - 1}: {e}")

finally:
    # 关闭浏览器
    driver.quit()
