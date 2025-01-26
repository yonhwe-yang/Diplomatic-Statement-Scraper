from selenium import webdriver
from selenium.webdriver.common.by import By
import re
from openpyxl import Workbook, load_workbook
import os
import time
import pandas as pd


# 读取 Excel 文件
df = pd.read_excel(r"C:\Users\86159\Desktop\外交语料\外交部\valid_pages.xlsx")

# 提取 'urls' 列并转换为列表
url_list = df['URL'].dropna().tolist()


names = ["林剑", "毛宁", "汪文斌", "赵立坚", "耿爽", "陆慷", "华春莹","郭嘉昆"]
#处理title
def process_title(text):
    match=re.match(r"(.+?)（(\d{4})-(\d{2})-(\d{2})）", text)
    if match:
        title = match.group(1).strip()  # 提取标题
        year, month, day = int(match.group(2)), int(match.group(3)), int(match.group(4))
        formatted_date = f"{year}/{month}/{day}"  # 格式化日期为要求的形式
        return {"标题": title, "时间": formatted_date}
    else:
        return {"标题": None, "时间": None}
#处理对话
def process_dialogue(question, answer):

    result = {
        "媒体": None,
        "提问": None,
        "发言人": None,
        "回答": None
    }

    # 处理提问部分
    if question:
        parts = re.split(r"[：:]", question, 1)#第一个：或:
        if len(parts) == 2:
            media_part, question_part = parts
            media = media_part.replace("记者", "").strip()  # 提取媒体名称
            result["媒体"] = media
            result["提问"] = question_part.strip()
        else:
            result["媒体"]="无"
            result["提问"]=question.strip()



    # 处理回答部分
    if answer:
        parts = re.split(r"[：:]", answer, 1)
        if len(parts) == 2:  # 确保分割后有两个部分
            spokesperson_part, answer_part = parts
            spokesperson_name = spokesperson_part.strip()
            # 检查spokesperson_part是否为指定的名字之一
            if spokesperson_name in names:
                result["发言人"] = spokesperson_name
                result["回答"] = answer_part.strip()
            elif spokesperson_name in ["答","回答","发言人"]:
                # 如果spokesperson_part不是指定的名字之一，则在title中寻找这些名字
                found_name = next((name for name in names if name in title), None)
                result["发言人"] = found_name if found_name else "答"
                result["回答"] = answer_part.strip()
            else:
                result["发言人"]="无"
                result["回答"] = answer.strip()
        else:
            # 如果没有分割出两个部分，将整个answer作为回答部分，发言人设置为标题中人名或“答”
            result["回答"] = answer.strip()
            found_name = next((name for name in names if name in title), None)
            result["发言人"] = found_name if found_name else "发言"

    return result

#输出函数
def append_to_excel(dialogues, link, title, date, strong_count,output_file="外交部记者会输出.xlsx"):
    # 定义 Excel 列名
    headers = ["序号", "链接", "标题", "日期", "媒体", "大主题", "小主题", "提问", "发言人", "回答"]

    # 检查文件是否已存在
    if not os.path.exists(output_file):
        # 如果文件不存在，则创建一个新的工作簿并写入标题
        wb = Workbook()
        ws = wb.active
        ws.title = "外交部记者会"
        ws.append(headers)  # 写入列名
        wb.save(output_file)

    # 加载已有的工作簿
    wb = load_workbook(output_file)
    ws = wb.active

    # 遍历对话并逐行写入
    for idx, (question, answer) in enumerate(dialogues, start=1):
        processed = process_dialogue(question, answer)  # 调用您已有的 process_dialogue 函数
        ws.append([
            strong_count,              # 序号
            link,                  # 链接
            title,                 # 标题
            date,                  # 日期
            processed["媒体"],      # 媒体
            "",                    # 大主题（空）
            "",                    # 小主题（空）
            processed["提问"],      # 提问
            processed["发言人"],    # 发言人
            processed["回答"]       # 回答
        ])


    # 保存文件
    wb.save(output_file)
    print(f"数据已写入到 {output_file}")

# 使用默认路径初始化 ChromeDriver
driver = webdriver.Chrome()

try:
    for url in url_list:
        # 打开目标网页
        '''url = "https://www.mfa.gov.cn/fyrbt_673021/dhdw_673027/index_5.shtml"'''
        driver.get(url)

        # 查找 div.newsBd 下的 <li> 中的 <a> 标签
        a_elements = driver.find_elements(By.CSS_SELECTOR, "div.newsBd li a")

        count = 0  # 初始化计数器，用于跟踪当前处理的链接

        while count < len(a_elements):
            # 获取当前链接与标题、时间
            element = a_elements[count]
            link = element.get_attribute("href")
            text = element.text.strip()
            # 处理文本中的标题和时间
            result = process_title(text)
            title = result['标题']
            date = result['时间']

            # 输出处理后的标题和时间
            print(f"标题: {title}, 时间: {date}, 链接: {link}")

            # 进入链接并等待页面加载
            driver.get(link)
            time.sleep(2)  # 添加适当的等待时间，确保页面加载完成

            # 获取 #News_Body_Txt_A 元素的文本内容
            try:
                news_body = driver.find_element(By.CSS_SELECTOR, "#News_Body_Txt_A")

                # 获取所有<p>标签
                paragraphs = news_body.find_elements(By.TAG_NAME, "p")

                dialogues = []  # 用于存储对话
                current_question = None  # 当前提问
                current_answers = []  # 当前回答（可能有多段）
                opening_answers = []  # 用于收集开头没有提问的普通文本
                strong_count = 0#check on strong
                # 遍历每个段落
                for paragraph in paragraphs:
                    text = driver.execute_script("return arguments[0].innerText;", paragraph)
                    text=text.strip()# 获取段落文本
                    if not text:  # 跳过空段落
                        continue

                    strong_elements = paragraph.find_elements(By.XPATH, ".//b | .//strong")
                    if strong_elements or "***" in text:
                        question = text
                        strong_count += 1
                        # 如果开头没有提问的段落存在，保存它们为独立对话
                        if opening_answers:
                            dialogues.append(("", "\n".join(opening_answers)))
                            opening_answers = []  # 清空开头段落

                        elif current_question:  # 如果已有提问，保存之前的对话
                            dialogues.append((current_question, "\n".join(current_answers)))


                        current_question = question  # 更新当前提问
                        current_answers = []  # 清空回答
                    else:
                        # 如果没有 <strong>，判断是开头没有提问的段落还是当前提问的回答
                        if current_question:
                            current_answers.append(text)  # 如果已有提问，将其归类为该提问的回答

                        else:
                            opening_answers.append(text)  # 如果没有提问，归类为开头段落

                    # 如果最后还有未保存的对话，保存它们
                if current_question:
                    dialogues.append((current_question, "\n".join(current_answers)))

                elif opening_answers:  # 如果只有开头的普通文本
                    dialogues.append(("", "\n".join(opening_answers)))

                # 输出对话内容
                print("识别到的对话：")
                for i, (question, answer) in enumerate(dialogues, 1):
                    processed = process_dialogue(question, answer)
                    print(f"对话 {i}:")
                    print(f"媒体: {processed['媒体']}")
                    print(f"提问: {processed['提问']}")
                    print(f"发言人: {processed['发言人']}")
                    print(f"回答: {processed['回答']}")
                    print("-" * 30)

                append_to_excel(dialogues, link, title, date,strong_count)

            except Exception as e:
                print(f"无法获取内容，错误信息：{e}")

            # 返回上一页，准备进入下一个链接
            driver.back()
            time.sleep(2)  # 确保返回到前一个页面后，页面完全加载

            # 获取当前页面中的所有链接（通过重新获取元素）
            a_elements = driver.find_elements(By.CSS_SELECTOR, "div.newsBd li a")

            # 增加计数器，进入下一个链接
            count += 1



finally:
    driver.quit()